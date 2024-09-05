import openpyxl
import click
from openpyxl.styles import PatternFill

@click.command()
@click.argument('inputfile')
@click.argument('outputfile')


def main(inputfile, outputfile):
    # Load the workbook
    input = openpyxl.load_workbook(inputfile)
    output = openpyxl.load_workbook(outputfile)
    print('Workbook loaded')
    wsin = input.active
    wsout = output.active
    print('Total number of rows: '+str(wsout.max_row)+'. And total number of columns: '+str(wsout.max_column))
    
    # update color of first 2 rows
    my_fill_1 = PatternFill('solid', start_color="00ffff")
    for y in range(1, wsout.max_column+1):
        wsout.cell(row=1, column=y).fill = my_fill_1
    my_fill_2 = PatternFill('solid', start_color="ffff00")
    for y in range(1, wsout.max_column+1):
        wsout.cell(row=2, column=y).fill = my_fill_2
    output.save(outputfile)

main()
